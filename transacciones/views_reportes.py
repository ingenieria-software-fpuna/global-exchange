"""
Vistas para reportes de transacciones y ganancias
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Sum, Count, Q, F, DecimalField
from django.db.models.functions import TruncDate, Coalesce
from django.utils import timezone
from datetime import datetime, timedelta, date
from decimal import Decimal
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from .models import Transaccion, TipoOperacion, EstadoTransaccion
from .forms import FiltroReporteForm
from monedas.models import Moneda
from tasa_cambio.models import TasaCambio
from monedas.templatetags.moneda_extras import moneda_format
from configuracion.models import ConfiguracionSistema





def aplicar_filtros_transacciones(queryset, form_data):
    """
    Aplica filtros al queryset de transacciones según los datos del formulario.
    """
    fecha_desde = form_data.get('fecha_desde')
    fecha_hasta = form_data.get('fecha_hasta')
    tipo_operacion = form_data.get('tipo_operacion')
    estado = form_data.get('estado')
    moneda = form_data.get('moneda')
    cliente = form_data.get('cliente')
    usuario = form_data.get('usuario')
    
    # Aplicar filtros
    if fecha_desde:
        queryset = queryset.filter(fecha_creacion__date__gte=fecha_desde)
    
    if fecha_hasta:
        queryset = queryset.filter(fecha_creacion__date__lte=fecha_hasta)
    
    if tipo_operacion:
        queryset = queryset.filter(tipo_operacion=tipo_operacion)
    
    if estado:
        queryset = queryset.filter(estado=estado)
    
    if moneda:
        queryset = queryset.filter(
            Q(moneda_origen=moneda) | Q(moneda_destino=moneda)
        )
    
    if cliente:
        queryset = queryset.filter(cliente=cliente)
    
    if usuario:
        queryset = queryset.filter(usuario=usuario)
    
    return queryset, fecha_desde, fecha_hasta


def calcular_ganancias_por_tipo(transacciones):
    """
    Calcula las ganancias por tipo de operación (Compra/Venta) en PYG.
    
    Args:
        transacciones: QuerySet de transacciones a procesar
        
    Returns:
        dict: Diccionario con estructura {codigo_tipo: {'tipo': TipoOperacion, 'ganancia_neta_pyg': Decimal, 
              'comisiones_pyg': Decimal, 'descuentos_pyg': Decimal, 'cantidad': int}}
    """
    ganancias_por_tipo = {}
    
    for tipo in TipoOperacion.objects.filter(activo=True):
        trans_tipo = transacciones.filter(tipo_operacion=tipo)
        
        if trans_tipo.exists():
            totales = trans_tipo.aggregate(
                total_comision=Coalesce(Sum('monto_comision'), Decimal('0.00')),
                total_descuento=Coalesce(Sum('monto_descuento'), Decimal('0.00'))
            )
            
            comisiones_pyg = totales['total_comision']
            descuentos_pyg = totales['total_descuento']
            ganancia_neta_pyg = comisiones_pyg - descuentos_pyg
            
            ganancias_por_tipo[tipo.codigo] = {
                'tipo': tipo,
                'comisiones_pyg': comisiones_pyg,
                'descuentos_pyg': descuentos_pyg,
                'ganancia_neta_pyg': ganancia_neta_pyg,
                'cantidad': trans_tipo.count(),
            }
    
    return ganancias_por_tipo


@login_required
@permission_required('transacciones.view_reporte_transacciones', raise_exception=True)
def reporte_transacciones(request):
    """
    Vista principal del reporte de transacciones con filtros.
    """
    form = FiltroReporteForm(request.GET or None)
    transacciones = Transaccion.objects.all().select_related(
        'tipo_operacion', 'estado', 'moneda_origen', 'moneda_destino',
        'cliente', 'usuario', 'tauser'
    )
    
    fecha_desde = None
    fecha_hasta = None
    
    if form.is_valid():
        transacciones, fecha_desde, fecha_hasta = aplicar_filtros_transacciones(
            transacciones, 
            form.cleaned_data
        )
    
    # Si no hay fechas especificadas, usar últimos 30 días para los gráficos
    if not fecha_desde:
        fecha_desde = timezone.now().date() - timedelta(days=30)
    if not fecha_hasta:
        fecha_hasta = timezone.now().date()
    
    # Ordenar por fecha descendente
    transacciones = transacciones.order_by('-fecha_creacion')
    
    # Calcular estadísticas
    total_canceladas_anuladas = transacciones.filter(
        estado__codigo__in=['CANCELADA', 'ANULADA']
    ).count()
    
    estadisticas = {
        'total': transacciones.count(),
        'pagadas': transacciones.filter(estado__codigo='PAGADA').count(),
        'pendientes': transacciones.filter(estado__codigo='PENDIENTE').count(),
        'canceladas': total_canceladas_anuladas,
        'compras': transacciones.filter(tipo_operacion__codigo='COMPRA').count(),
        'ventas': transacciones.filter(tipo_operacion__codigo='VENTA').count(),
    }
    
    # Preparar datos para el gráfico de líneas (transacciones por día)
    import json
    transacciones_por_dia = transacciones.annotate(
        fecha=TruncDate('fecha_creacion')
    ).values('fecha').annotate(
        cantidad=Count('id')
    ).order_by('fecha')
    
    # Generar todas las fechas del rango
    fechas_grafico = []
    cantidades_grafico = []
    fecha_actual = fecha_desde
    while fecha_actual <= fecha_hasta:
        fechas_grafico.append(fecha_actual.strftime('%Y-%m-%d'))
        # Buscar si hay transacciones en esa fecha
        cantidad = 0
        for item in transacciones_por_dia:
            if item['fecha'] == fecha_actual:
                cantidad = item['cantidad']
                break
        cantidades_grafico.append(cantidad)
        fecha_actual += timedelta(days=1)
    
    datos_grafico = json.dumps({
        'fechas': fechas_grafico,
        'cantidades': cantidades_grafico
    })
    
    # Implementar paginación
    paginator = Paginator(transacciones, 20)  # 20 transacciones por página
    page = request.GET.get('page', 1)
    
    try:
        transacciones_paginadas = paginator.page(page)
    except PageNotAnInteger:
        transacciones_paginadas = paginator.page(1)
    except EmptyPage:
        transacciones_paginadas = paginator.page(paginator.num_pages)
    
    # Calcular ganancia en PYG para cada transacción
    for transaccion in transacciones_paginadas:
        # TANTO LA COMISIÓN COMO EL DESCUENTO SIEMPRE ESTÁN EN PYG (por diseño del sistema)
        # En COMPRA: moneda_origen=PYG, comisión/descuento en PYG
        # En VENTA: moneda_destino=PYG, comisión/descuento en PYG
        transaccion.comision_pyg = transaccion.monto_comision
        transaccion.descuento_pyg = transaccion.monto_descuento
        
        # Solo calcular ganancia para transacciones completadas
        if transaccion.estado.codigo in ['PAGADA', 'ENTREGADA', 'RETIRADO']:
            transaccion.ganancia_pyg = transaccion.comision_pyg - transaccion.descuento_pyg
        else:
            # Transacciones canceladas/anuladas/pendientes no tienen ganancia
            transaccion.ganancia_pyg = Decimal('0.00')
    
    context = {
        'form': form,
        'transacciones': transacciones_paginadas,
        'estadisticas': estadisticas,
        'datos_grafico': datos_grafico,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'total_registros': transacciones.count(),
        'page_obj': transacciones_paginadas,
        'is_paginated': paginator.num_pages > 1,
    }
    
    return render(request, 'transacciones/reporte_transacciones.html', context)


@login_required
@permission_required('transacciones.view_reporte_ganancias', raise_exception=True)
def reporte_ganancias(request):
    """
    Vista principal del tablero de control de ganancias con filtros.
    """
    form = FiltroReporteForm(request.GET or None)
    
    # Solo transacciones pagadas generan ganancias
    transacciones = Transaccion.objects.filter(
        estado__codigo__in=['PAGADA', 'ENTREGADA', 'RETIRADO']
    ).select_related(
        'tipo_operacion', 'estado', 'moneda_origen', 'moneda_destino'
    )
    
    fecha_desde = None
    fecha_hasta = None
    
    if form.is_valid():
        transacciones, fecha_desde, fecha_hasta = aplicar_filtros_transacciones(
            transacciones, 
            form.cleaned_data
        )
    
    # Calcular ganancias por moneda
    ganancias_por_moneda = {}
    
    # Calcular totales UNA SOLA VEZ sobre todas las transacciones
    # (no por moneda para evitar contar transacciones dos veces)
    totales_generales = transacciones.aggregate(
        total_comision=Coalesce(Sum('monto_comision'), Decimal('0.00')),
        total_descuento=Coalesce(Sum('monto_descuento'), Decimal('0.00'))
    )
    
    total_comisiones_pyg = totales_generales['total_comision']
    total_descuentos_pyg = totales_generales['total_descuento']
    total_ganancia_pyg = total_comisiones_pyg - total_descuentos_pyg
    
    # Obtener configuración del sistema
    try:
        config = ConfiguracionSistema.obtener_configuracion()
        moneda_base = config.moneda_base
    except:
        moneda_base = Moneda.objects.filter(codigo='PYG').first()
    
    for moneda in Moneda.objects.filter(es_activa=True):
        # Filtrar transacciones donde la moneda es origen O destino (igual que reporte de transacciones)
        transacciones_moneda = transacciones.filter(
            Q(moneda_origen=moneda) | Q(moneda_destino=moneda)
        )
        
        if transacciones_moneda.count() == 0:
            continue
        
        # Calcular conversión a PYG sumando cada transacción
        # TANTO LA COMISIÓN COMO EL DESCUENTO SIEMPRE ESTÁN EN PYG (por diseño del sistema)
        comision_pyg = Decimal('0.00')
        descuento_pyg = Decimal('0.00')
        
        for trans in transacciones_moneda:
            comision_pyg += trans.monto_comision
            descuento_pyg += trans.monto_descuento
        
        ganancia_pyg = comision_pyg - descuento_pyg
        
        if ganancia_pyg != Decimal('0.00'):  # Mostrar incluso si es negativo
            ganancias_por_moneda[moneda.codigo] = {
                'moneda': moneda,
                'total_comision': None,  # No aplica porque mezclamos origen/destino
                'total_descuento': None,  # No aplica porque mezclamos origen/destino
                'ganancia_neta': None,   # No aplica porque mezclamos origen/destino
                'comision_pyg': comision_pyg,
                'descuento_pyg': descuento_pyg,
                'ganancia_pyg': ganancia_pyg,
                'cantidad_transacciones': transacciones_moneda.count(),
            }
            # NO sumamos a los totales aquí porque ya se calcularon arriba
            # (cada transacción aparece en múltiples monedas, causaría duplicación)
    
    # Estadísticas generales
    estadisticas = {
        'total_transacciones': transacciones.count(),
        'total_comisiones_pyg': total_comisiones_pyg,
        'total_descuentos_pyg': total_descuentos_pyg,
    }
    
    # Desglose por tipo de operación (convertido a PYG)
    ganancias_por_tipo = calcular_ganancias_por_tipo(transacciones)
    
    context = {
        'form': form,
        'ganancias_por_moneda': ganancias_por_moneda,
        'ganancias_por_tipo': ganancias_por_tipo,
        'estadisticas': estadisticas,
        'total_ganancia_pyg': total_ganancia_pyg,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'moneda_base': moneda_base,
    }
    
    return render(request, 'transacciones/reporte_ganancias.html', context)


@login_required
@permission_required('transacciones.view_reporte_transacciones', raise_exception=True)
def exportar_transacciones_excel(request):
    """
    Exporta el reporte de transacciones a Excel.
    """
    form = FiltroReporteForm(request.GET or None)
    transacciones = Transaccion.objects.all().select_related(
        'tipo_operacion', 'estado', 'moneda_origen', 'moneda_destino',
        'cliente', 'usuario'
    )
    
    if form.is_valid():
        transacciones, fecha_desde, fecha_hasta = aplicar_filtros_transacciones(
            transacciones, 
            form.cleaned_data
        )
    
    transacciones = transacciones.order_by('-fecha_creacion')
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Transacciones"
    
    # Estilos
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'ID Transacción', 'Fecha', 'Usuario', 'Cliente', 'Tipo', 'Estado',
        'Moneda Origen', 'Monto Origen', 'Moneda Destino', 'Monto Destino',
        'Tasa Cambio', 'Requiere Retiro', 'TAUSER', 'Comisión (PYG)', 'Descuento (PYG)', 'Ganancia (PYG)'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row_num, trans in enumerate(transacciones, 2):
        # TANTO LA COMISIÓN COMO EL DESCUENTO SIEMPRE ESTÁN EN PYG
        comision_pyg = trans.monto_comision
        descuento_pyg = trans.monto_descuento
        
        # Calcular ganancia en PYG solo para transacciones completadas
        if trans.estado.codigo in ['PAGADA', 'ENTREGADA', 'RETIRADO']:
            ganancia_pyg = comision_pyg - descuento_pyg
        else:
            ganancia_pyg = Decimal('0.00')
        
        ws.cell(row=row_num, column=1).value = trans.id_transaccion
        ws.cell(row=row_num, column=2).value = trans.fecha_creacion.strftime('%Y-%m-%d %H:%M')
        ws.cell(row=row_num, column=3).value = f"{trans.usuario.nombre} {trans.usuario.apellido}".strip() if trans.usuario else '-'
        ws.cell(row=row_num, column=4).value = trans.cliente.nombre_comercial if trans.cliente else 'Cliente Casual'
        ws.cell(row=row_num, column=5).value = trans.tipo_operacion.nombre
        ws.cell(row=row_num, column=6).value = trans.estado.nombre
        ws.cell(row=row_num, column=7).value = trans.moneda_origen.codigo
        ws.cell(row=row_num, column=8).value = float(trans.monto_origen)
        ws.cell(row=row_num, column=9).value = trans.moneda_destino.codigo
        ws.cell(row=row_num, column=10).value = float(trans.monto_destino)
        ws.cell(row=row_num, column=11).value = float(trans.tasa_cambio)
        ws.cell(row=row_num, column=12).value = 'Sí' if trans.requiere_retiro_fisico else 'No'
        ws.cell(row=row_num, column=13).value = trans.tauser.nombre if trans.tauser else '-'
        ws.cell(row=row_num, column=14).value = float(comision_pyg)
        ws.cell(row=row_num, column=15).value = float(descuento_pyg)
        ws.cell(row=row_num, column=16).value = float(ganancia_pyg)
        
        # Aplicar bordes
        for col_num in range(1, 17):
            ws.cell(row=row_num, column=col_num).border = border
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Preparar respuesta
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_transacciones_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


@login_required
@permission_required('transacciones.view_reporte_transacciones', raise_exception=True)
def exportar_transacciones_pdf(request):
    """
    Exporta el reporte de transacciones a PDF.
    """
    form = FiltroReporteForm(request.GET or None)
    transacciones = Transaccion.objects.all().select_related(
        'tipo_operacion', 'estado', 'moneda_origen', 'moneda_destino',
        'cliente', 'usuario'
    )
    
    fecha_desde = None
    fecha_hasta = None
    
    if form.is_valid():
        transacciones, fecha_desde, fecha_hasta = aplicar_filtros_transacciones(
            transacciones, 
            form.cleaned_data
        )
    
    transacciones = transacciones.order_by('-fecha_creacion')
    
    # Calcular estadísticas para el PDF
    total_canceladas_anuladas = transacciones.filter(
        estado__codigo__in=['CANCELADA', 'ANULADA']
    ).count()
    
    estadisticas = {
        'total': transacciones.count(),
        'pagadas': transacciones.filter(estado__codigo='PAGADA').count(),
        'pendientes': transacciones.filter(estado__codigo='PENDIENTE').count(),
        'canceladas': total_canceladas_anuladas,
        'compras': transacciones.filter(tipo_operacion__codigo='COMPRA').count(),
        'ventas': transacciones.filter(tipo_operacion__codigo='VENTA').count(),
    }
    
    # Crear PDF en orientación horizontal para que quepa el UUID completo
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0d6efd'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Título
    elements.append(Paragraph("Reporte de Transacciones", title_style))
    elements.append(Spacer(1, 12))
    
    # Período
    if fecha_desde and fecha_hasta:
        periodo_text = f"Período: {fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
        elements.append(Paragraph(periodo_text, styles['Normal']))
        elements.append(Spacer(1, 12))
    
    # Estadísticas
    stats_text = (
        f"<b>Total de transacciones:</b> {estadisticas['total']} | "
        f"<b>Pagadas:</b> {estadisticas['pagadas']} | "
        f"<b>Pendientes:</b> {estadisticas['pendientes']} | "
        f"<b>Canceladas:</b> {estadisticas['canceladas']}"
    )
    stats_para = Paragraph(stats_text, styles['Normal'])
    elements.append(stats_para)
    elements.append(Spacer(1, 20))
    
    # Datos de la tabla
    data = [['ID', 'Fecha', 'Usuario', 'Cliente', 'Tipo', 'Estado', 'Monto origen', 'Monto destino', 'Comisión (PYG)', 'Descuento (PYG)', 'Ganancia (PYG)']]
    
    for trans in transacciones[:100]:  # Limitar a 100 registros en PDF
        usuario_nombre = f"{trans.usuario.nombre} {trans.usuario.apellido}".strip() if trans.usuario else '-'
        cliente_nombre = trans.cliente.nombre_comercial if trans.cliente else 'Casual'
        
        # Formatear montos usando el filtro moneda_format
        monto_origen_fmt = moneda_format(trans.monto_origen, trans.moneda_origen.codigo)
        monto_destino_fmt = moneda_format(trans.monto_destino, trans.moneda_destino.codigo)
        
        # TANTO LA COMISIÓN COMO EL DESCUENTO SIEMPRE ESTÁN EN PYG
        comision_pyg = trans.monto_comision
        descuento_pyg = trans.monto_descuento
        
        comision_pyg_fmt = moneda_format(comision_pyg, 'PYG')
        descuento_pyg_fmt = moneda_format(descuento_pyg, 'PYG')
        
        # Calcular ganancia en PYG solo para transacciones completadas
        if trans.estado.codigo in ['PAGADA', 'ENTREGADA', 'RETIRADO']:
            ganancia_pyg = comision_pyg - descuento_pyg
        else:
            ganancia_pyg = Decimal('0.00')
        ganancia_pyg_fmt = moneda_format(ganancia_pyg, 'PYG')
        
        data.append([
            trans.id_transaccion,  # ID completo sin cortar
            trans.fecha_creacion.strftime('%d/%m/%Y'),
            usuario_nombre[:15],
            cliente_nombre[:15],
            trans.tipo_operacion.codigo,
            trans.estado.codigo,
            monto_origen_fmt,
            monto_destino_fmt,
            comision_pyg_fmt,
            descuento_pyg_fmt,
            ganancia_pyg_fmt
        ])
    
    # Crear tabla con anchos de columna específicos
    # Ancho total de página A4 horizontal: aproximadamente 800 puntos disponibles
    col_widths = [
        100,  # ID (UUID - más reducido)
        55,   # Fecha
        70,   # Usuario
        70,   # Cliente
        45,   # Tipo
        50,   # Estado
        75,   # Origen
        75,   # Destino
        70,   # Comisión (PYG)
        70,   # Descuento (PYG)
        70,   # Ganancia (PYG)
    ]
    
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('WORDWRAP', (0, 0), (-1, -1), True),  # Permitir ajuste de texto
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_transacciones_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response


@login_required
@permission_required('transacciones.view_reporte_ganancias', raise_exception=True)
def exportar_ganancias_excel(request):
    """
    Exporta el tablero de control de ganancias a Excel.
    """
    form = FiltroReporteForm(request.GET or None)
    transacciones = Transaccion.objects.filter(
        estado__codigo__in=['PAGADA', 'ENTREGADA', 'RETIRADO']
    ).select_related('tipo_operacion', 'moneda_origen', 'moneda_destino')
    
    fecha_desde = None
    fecha_hasta = None
    
    if form.is_valid():
        transacciones, fecha_desde, fecha_hasta = aplicar_filtros_transacciones(
            transacciones, 
            form.cleaned_data
        )
    
    # Calcular estadísticas (igual que en la vista principal)
    total_ganancia_pyg = Decimal('0.00')
    total_comisiones_pyg = Decimal('0.00')
    total_descuentos_pyg = Decimal('0.00')
    
    # Calcular por tipo de operación (reutilizando función auxiliar)
    ganancias_por_tipo = calcular_ganancias_por_tipo(transacciones)
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Ganancias"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    total_font = Font(bold=True)
    
    # Título
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "TABLERO DE CONTROL DE GANANCIAS"
    title_cell.font = Font(size=14, bold=True, color="366092")
    title_cell.alignment = Alignment(horizontal='center')
    
    # Período si existe
    current_row = 2
    if fecha_desde and fecha_hasta:
        ws.merge_cells(f'A{current_row}:E{current_row}')
        periodo_cell = ws.cell(row=current_row, column=1)
        periodo_cell.value = f"Período: {fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
        periodo_cell.alignment = Alignment(horizontal='center')
        current_row += 1
    
    current_row += 1
    
    # Encabezados por moneda
    ws.cell(row=current_row, column=1).value = "Moneda"
    ws.cell(row=current_row, column=2).value = "Total Comisiones"
    ws.cell(row=current_row, column=3).value = "Total Descuentos"
    ws.cell(row=current_row, column=4).value = "Ganancia Neta"
    ws.cell(row=current_row, column=5).value = "Ganancia en PYG"
    
    for col_num in range(1, 6):
        cell = ws.cell(row=current_row, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Calcular ganancias por moneda
    row_num = current_row + 1
    
    for moneda in Moneda.objects.filter(es_activa=True):
        transacciones_moneda = transacciones.filter(moneda_origen=moneda)
        
        ganancia = transacciones_moneda.aggregate(
            total_comision=Coalesce(Sum('monto_comision'), Decimal('0.00')),
            total_descuento=Coalesce(Sum('monto_descuento'), Decimal('0.00'))
        )
        
        ganancia_neta = ganancia['total_comision'] - ganancia['total_descuento']
        
        if ganancia_neta > 0:
            # Convertir a PYG
            if moneda.codigo == 'PYG':
                ganancia_pyg = ganancia_neta
                comision_pyg = ganancia['total_comision']
                descuento_pyg = ganancia['total_descuento']
            else:
                tasa_promedio = transacciones_moneda.aggregate(
                    tasa_avg=Coalesce(Sum('tasa_cambio') / Count('id'), Decimal('1.00'))
                )['tasa_avg']
                ganancia_pyg = ganancia_neta * tasa_promedio
                comision_pyg = ganancia['total_comision'] * tasa_promedio
                descuento_pyg = ganancia['total_descuento'] * tasa_promedio
            
            ws.cell(row=row_num, column=1).value = moneda.codigo
            ws.cell(row=row_num, column=2).value = float(ganancia['total_comision'])
            ws.cell(row=row_num, column=3).value = float(ganancia['total_descuento'])
            ws.cell(row=row_num, column=4).value = float(ganancia_neta)
            ws.cell(row=row_num, column=5).value = float(ganancia_pyg)
            
            total_ganancia_pyg += ganancia_pyg
            total_comisiones_pyg += comision_pyg
            total_descuentos_pyg += descuento_pyg
            row_num += 1
    
    # Fila de total
    ws.cell(row=row_num, column=1).value = "TOTAL (PYG)"
    ws.cell(row=row_num, column=2).value = float(total_comisiones_pyg)
    ws.cell(row=row_num, column=3).value = float(total_descuentos_pyg)
    ws.cell(row=row_num, column=4).value = ""
    ws.cell(row=row_num, column=5).value = float(total_ganancia_pyg)
    
    for col_num in range(1, 6):
        cell = ws.cell(row=row_num, column=col_num)
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal='center')
    
    row_num += 2
    
    # Estadísticas generales
    ws.merge_cells(f'A{row_num}:E{row_num}')
    stats_cell = ws.cell(row=row_num, column=1)
    stats_cell.value = "ESTADÍSTICAS GENERALES"
    stats_cell.font = Font(size=12, bold=True)
    stats_cell.alignment = Alignment(horizontal='center')
    row_num += 1
    
    stats_data = [
        ['Total Transacciones', transacciones.count()],
        ['Total Comisiones (PYG)', float(total_comisiones_pyg)],
        ['Total Descuentos (PYG)', float(total_descuentos_pyg)],
        ['Ganancia Total (PYG)', float(total_ganancia_pyg)],
    ]
    
    for stat in stats_data:
        ws.cell(row=row_num, column=1).value = stat[0]
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=2).value = stat[1]
        row_num += 1
    
    # Ganancias por tipo de operación
    if ganancias_por_tipo:
        row_num += 1
        ws.merge_cells(f'A{row_num}:E{row_num}')
        tipo_cell = ws.cell(row=row_num, column=1)
        tipo_cell.value = "GANANCIAS POR TIPO DE OPERACIÓN"
        tipo_cell.font = Font(size=12, bold=True)
        tipo_cell.alignment = Alignment(horizontal='center')
        row_num += 1
        
        for codigo, data in ganancias_por_tipo.items():
            ws.cell(row=row_num, column=1).value = data['tipo'].nombre
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=2).value = float(data['ganancia_neta_pyg'])
            ws.cell(row=row_num, column=3).value = data['cantidad']
            row_num += 1
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    
    # Preparar respuesta
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_ganancias_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


@login_required
@permission_required('transacciones.view_reporte_ganancias', raise_exception=True)
def exportar_ganancias_pdf(request):
    """
    Exporta el tablero de control de ganancias a PDF.
    """
    form = FiltroReporteForm(request.GET or None)
    transacciones = Transaccion.objects.filter(
        estado__codigo__in=['PAGADA', 'ENTREGADA', 'RETIRADO']
    ).select_related('tipo_operacion', 'moneda_origen', 'moneda_destino')
    
    fecha_desde = None
    fecha_hasta = None
    
    if form.is_valid():
        transacciones, fecha_desde, fecha_hasta = aplicar_filtros_transacciones(
            transacciones, 
            form.cleaned_data
        )
    
    # Calcular estadísticas (igual que en la vista)
    total_ganancia_pyg = Decimal('0.00')
    total_comisiones_pyg = Decimal('0.00')
    total_descuentos_pyg = Decimal('0.00')
    
    # Calcular por tipo de operación (reutilizando función auxiliar)
    ganancias_por_tipo = calcular_ganancias_por_tipo(transacciones)
    
    # Crear PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0d6efd'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Título
    elements.append(Paragraph("Tablero de Control de Ganancias", title_style))
    elements.append(Spacer(1, 12))
    
    # Período
    if fecha_desde and fecha_hasta:
        periodo_text = f"Período: {fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
        elements.append(Paragraph(periodo_text, styles['Normal']))
        elements.append(Spacer(1, 12))
    
    # Calcular ganancias
    data = [['Moneda', 'Comisiones', 'Descuentos', 'Ganancia Neta', 'Ganancia PYG']]
    
    for moneda in Moneda.objects.filter(es_activa=True):
        transacciones_moneda = transacciones.filter(moneda_origen=moneda)
        
        ganancia = transacciones_moneda.aggregate(
            total_comision=Coalesce(Sum('monto_comision'), Decimal('0.00')),
            total_descuento=Coalesce(Sum('monto_descuento'), Decimal('0.00'))
        )
        
        ganancia_neta = ganancia['total_comision'] - ganancia['total_descuento']
        
        if ganancia_neta > 0:
            if moneda.codigo == 'PYG':
                ganancia_pyg = ganancia_neta
                comision_pyg = ganancia['total_comision']
                descuento_pyg = ganancia['total_descuento']
            else:
                tasa_promedio = transacciones_moneda.aggregate(
                    tasa_avg=Coalesce(Sum('tasa_cambio') / Count('id'), Decimal('1.00'))
                )['tasa_avg']
                ganancia_pyg = ganancia_neta * tasa_promedio
                comision_pyg = ganancia['total_comision'] * tasa_promedio
                descuento_pyg = ganancia['total_descuento'] * tasa_promedio
            
            data.append([
                moneda.codigo,
                f"{ganancia['total_comision']:,.2f}",
                f"{ganancia['total_descuento']:,.2f}",
                f"{ganancia_neta:,.2f}",
                f"{ganancia_pyg:,.2f}"
            ])
            
            total_ganancia_pyg += ganancia_pyg
            total_comisiones_pyg += comision_pyg
            total_descuentos_pyg += descuento_pyg
    
    # Fila de total
    data.append([
        'TOTAL (PYG)',
        f"{total_comisiones_pyg:,.0f}",
        f"{total_descuentos_pyg:,.0f}",
        '',
        f"{total_ganancia_pyg:,.2f}"
    ])
    
    # Crear tabla de ganancias por moneda
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFD700')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 30))
    
    # Estadísticas generales
    elements.append(Paragraph("Estadísticas Generales", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    stats_data = [
        ['Métrica', 'Valor'],
        ['Total Transacciones', str(transacciones.count())],
        ['Total Comisiones (PYG)', f"{total_comisiones_pyg:,.0f}"],
        ['Total Descuentos (PYG)', f"{total_descuentos_pyg:,.0f}"],
        ['Ganancia Total (PYG)', f"{total_ganancia_pyg:,.0f}"]
    ]
    
    stats_table = Table(stats_data)
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(stats_table)
    
    # Ganancias por tipo de operación
    if ganancias_por_tipo:
        elements.append(Spacer(1, 30))
        elements.append(Paragraph("Ganancias por Tipo de Operación", styles['Heading2']))
        elements.append(Spacer(1, 12))
        
        tipo_data = [['Tipo de Operación', 'Ganancia Neta (PYG)', 'Transacciones']]
        for codigo, data_tipo in ganancias_por_tipo.items():
            tipo_data.append([
                data_tipo['tipo'].nombre,
                f"{data_tipo['ganancia_neta_pyg']:,.0f}",
                str(data_tipo['cantidad'])
            ])
        
        tipo_table = Table(tipo_data)
        tipo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(tipo_table)
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_ganancias_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response
